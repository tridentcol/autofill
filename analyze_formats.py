#!/usr/bin/env python3
"""
Script para analizar la estructura de los formatos Excel
"""
import openpyxl
from openpyxl import load_workbook
import xlrd
import json
import sys
from pathlib import Path

def analyze_xlsx(filepath):
    """Analiza archivos .xlsx"""
    print(f"\n{'='*80}")
    print(f"ANALIZANDO: {filepath}")
    print(f"{'='*80}\n")

    wb = load_workbook(filepath, data_only=False)

    analysis = {
        'filename': Path(filepath).name,
        'format': 'xlsx',
        'sheets': []
    }

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        print(f"\n📄 HOJA: {sheet_name}")
        print(f"   Dimensiones: {ws.dimensions}")
        print(f"   Max fila: {ws.max_row}, Max columna: {ws.max_column}")

        sheet_info = {
            'name': sheet_name,
            'dimensions': ws.dimensions,
            'max_row': ws.max_row,
            'max_column': ws.max_column,
            'merged_cells': [],
            'filled_cells': [],
            'empty_cells': [],
            'headers': [],
            'potential_fields': []
        }

        # Celdas combinadas
        if ws.merged_cells:
            print(f"\n   🔗 Celdas combinadas: {len(ws.merged_cells.ranges)}")
            for merged_range in ws.merged_cells.ranges:
                sheet_info['merged_cells'].append(str(merged_range))
                print(f"      - {merged_range}")

        # Analizar contenido (primeras 50 filas para no sobrecargar)
        print(f"\n   📋 Análisis de contenido (primeras 50 filas):")
        rows_to_check = min(50, ws.max_row)

        for row_idx in range(1, rows_to_check + 1):
            for col_idx in range(1, min(20, ws.max_column + 1)):  # Max 20 columnas
                cell = ws.cell(row=row_idx, column=col_idx)
                cell_ref = cell.coordinate

                if cell.value is not None and str(cell.value).strip():
                    cell_info = {
                        'ref': cell_ref,
                        'value': str(cell.value)[:100],  # Limitar tamaño
                        'row': row_idx,
                        'col': col_idx
                    }
                    sheet_info['filled_cells'].append(cell_info)

                    # Detectar posibles encabezados (primeras 5 filas con texto)
                    if row_idx <= 5 and isinstance(cell.value, str):
                        if any(keyword in cell.value.lower() for keyword in
                               ['nombre', 'fecha', 'firma', 'observ', 'item', 'descripcion',
                                'responsable', 'cargo', 'hora', 'proyecto', 'actividad',
                                'equipo', 'herramienta', 'estado', 'si', 'no', 'n/a']):
                            sheet_info['headers'].append(cell_info)

                else:
                    # Celdas vacías que podrían ser campos a llenar
                    # Solo guardamos algunas para no sobrecargar
                    if row_idx <= 30 and len(sheet_info['empty_cells']) < 100:
                        sheet_info['empty_cells'].append({
                            'ref': cell_ref,
                            'row': row_idx,
                            'col': col_idx
                        })

        print(f"      - Celdas con contenido: {len(sheet_info['filled_cells'])}")
        print(f"      - Celdas vacías detectadas: {len(sheet_info['empty_cells'])}")
        print(f"      - Posibles encabezados: {len(sheet_info['headers'])}")

        if sheet_info['headers']:
            print(f"\n   🏷️  Encabezados detectados:")
            for header in sheet_info['headers'][:10]:  # Mostrar primeros 10
                print(f"      - {header['ref']}: {header['value']}")

        analysis['sheets'].append(sheet_info)

    return analysis

def analyze_xls(filepath):
    """Analiza archivos .xls (formato antiguo)"""
    print(f"\n{'='*80}")
    print(f"ANALIZANDO: {filepath}")
    print(f"{'='*80}\n")

    wb = xlrd.open_workbook(filepath, formatting_info=True)

    analysis = {
        'filename': Path(filepath).name,
        'format': 'xls',
        'sheets': []
    }

    for sheet_idx in range(wb.nsheets):
        sheet = wb.sheet_by_index(sheet_idx)
        sheet_name = sheet.name

        print(f"\n📄 HOJA: {sheet_name}")
        print(f"   Filas: {sheet.nrows}, Columnas: {sheet.ncols}")

        sheet_info = {
            'name': sheet_name,
            'nrows': sheet.nrows,
            'ncols': sheet.ncols,
            'filled_cells': [],
            'empty_cells': [],
            'headers': [],
        }

        # Analizar contenido (primeras 50 filas)
        rows_to_check = min(50, sheet.nrows)

        for row_idx in range(rows_to_check):
            for col_idx in range(min(20, sheet.ncols)):
                cell = sheet.cell(row_idx, col_idx)
                cell_ref = f"{chr(65 + col_idx)}{row_idx + 1}"

                if cell.value:
                    cell_info = {
                        'ref': cell_ref,
                        'value': str(cell.value)[:100],
                        'row': row_idx + 1,
                        'col': col_idx + 1,
                        'type': cell.ctype
                    }
                    sheet_info['filled_cells'].append(cell_info)

                    # Detectar encabezados
                    if row_idx < 5 and cell.ctype == 1:  # Texto
                        if any(keyword in str(cell.value).lower() for keyword in
                               ['nombre', 'fecha', 'firma', 'observ', 'item', 'descripcion',
                                'responsable', 'cargo', 'hora', 'proyecto', 'actividad',
                                'equipo', 'herramienta', 'estado', 'si', 'no', 'n/a']):
                            sheet_info['headers'].append(cell_info)
                else:
                    if row_idx < 30 and len(sheet_info['empty_cells']) < 100:
                        sheet_info['empty_cells'].append({
                            'ref': cell_ref,
                            'row': row_idx + 1,
                            'col': col_idx + 1
                        })

        print(f"      - Celdas con contenido: {len(sheet_info['filled_cells'])}")
        print(f"      - Celdas vacías detectadas: {len(sheet_info['empty_cells'])}")
        print(f"      - Posibles encabezados: {len(sheet_info['headers'])}")

        if sheet_info['headers']:
            print(f"\n   🏷️  Encabezados detectados:")
            for header in sheet_info['headers'][:10]:
                print(f"      - {header['ref']}: {header['value']}")

        analysis['sheets'].append(sheet_info)

    return analysis

def main():
    files = [
        'INSPECCION VEHICULO CAMIONETA (4).xlsx',
        'PERMISO DE TRABAJO (8).xls',
        'INSPECCION HERRAMIENTAS Y O EQUIPOS (9).xlsx',
        'ANALISIS DE TRABAJO SEGURO (ATS) actual (15) (9).xls',
        'INSPECCION CAMION GRUA MANLIFT (15).xlsx'
    ]

    all_analysis = []

    for filepath in files:
        try:
            if filepath.endswith('.xlsx'):
                analysis = analyze_xlsx(filepath)
            else:
                analysis = analyze_xls(filepath)
            all_analysis.append(analysis)
        except Exception as e:
            print(f"\n❌ Error analizando {filepath}: {str(e)}")
            import traceback
            traceback.print_exc()

    # Guardar análisis completo
    with open('analysis_results.json', 'w', encoding='utf-8') as f:
        json.dump(all_analysis, f, indent=2, ensure_ascii=False)

    print(f"\n\n{'='*80}")
    print("✅ Análisis completado. Resultados guardados en: analysis_results.json")
    print(f"{'='*80}\n")

    # Resumen
    print("\n📊 RESUMEN GENERAL:")
    print(f"   Total de formatos analizados: {len(all_analysis)}")
    for analysis in all_analysis:
        print(f"\n   📄 {analysis['filename']}")
        print(f"      - Formato: {analysis['format']}")
        print(f"      - Hojas: {len(analysis['sheets'])}")
        for sheet in analysis['sheets']:
            print(f"         • {sheet['name']}: {len(sheet['headers'])} encabezados, {len(sheet['filled_cells'])} celdas con contenido")

if __name__ == '__main__':
    main()
