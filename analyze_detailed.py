#!/usr/bin/env python3
"""
Análisis detallado de formatos - muestra contenido de celdas
"""
import openpyxl
from openpyxl import load_workbook
import xlrd

def print_cells_content(filepath, max_rows=60):
    """Muestra el contenido de las celdas para entender la estructura"""
    print(f"\n{'='*100}")
    print(f"ANÁLISIS DETALLADO: {filepath}")
    print(f"{'='*100}\n")

    if filepath.endswith('.xlsx'):
        wb = load_workbook(filepath, data_only=False)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"\n{'='*100}")
            print(f"HOJA: {sheet_name}")
            print(f"{'='*100}\n")

            rows_to_show = min(max_rows, ws.max_row)
            cols_to_show = min(20, ws.max_column)

            # Crear una matriz para visualización
            for row_idx in range(1, rows_to_show + 1):
                row_content = []
                has_content = False
                for col_idx in range(1, cols_to_show + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    value = cell.value
                    if value is not None and str(value).strip():
                        has_content = True
                        # Truncar valores largos
                        value_str = str(value)[:40]
                        row_content.append(f"{cell.coordinate}: {value_str}")

                if has_content:
                    print(f"Fila {row_idx:3d}: {' | '.join(row_content)}")

    else:  # .xls
        wb = xlrd.open_workbook(filepath, formatting_info=True)
        for sheet_idx in range(wb.nsheets):
            sheet = wb.sheet_by_index(sheet_idx)
            sheet_name = sheet.name

            if sheet.nrows == 0:
                continue

            print(f"\n{'='*100}")
            print(f"HOJA: {sheet_name}")
            print(f"{'='*100}\n")

            rows_to_show = min(max_rows, sheet.nrows)
            cols_to_show = min(20, sheet.ncols)

            for row_idx in range(rows_to_show):
                row_content = []
                has_content = False
                for col_idx in range(cols_to_show):
                    cell = sheet.cell(row_idx, col_idx)
                    value = cell.value
                    if value:
                        has_content = True
                        value_str = str(value)[:40]
                        col_letter = chr(65 + col_idx) if col_idx < 26 else f"A{chr(65 + col_idx - 26)}"
                        cell_ref = f"{col_letter}{row_idx + 1}"
                        row_content.append(f"{cell_ref}: {value_str}")

                if has_content:
                    print(f"Fila {row_idx + 1:3d}: {' | '.join(row_content)}")

def main():
    files = [
        'INSPECCION VEHICULO CAMIONETA (4).xlsx',
        'PERMISO DE TRABAJO (8).xls',
        'INSPECCION HERRAMIENTAS Y O EQUIPOS (9).xlsx',
        'ANALISIS DE TRABAJO SEGURO (ATS) actual (15) (9).xls',
        'INSPECCION CAMION GRUA MANLIFT (15).xlsx'
    ]

    for filepath in files:
        try:
            print_cells_content(filepath)
        except Exception as e:
            print(f"\n❌ Error analizando {filepath}: {str(e)}")
            import traceback
            traceback.print_exc()

if __name__ == '__main__':
    main()
